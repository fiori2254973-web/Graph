from __future__ import annotations

import sys
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


class ReferenceCorpusTests(unittest.TestCase):
    def test_reference_corpus_contains_twelve_sdd_cases(self) -> None:
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl non installato")

        from examples import create_reference_workbooks

        original_file = Path(create_reference_workbooks.__file__).resolve().parent / "reference_cases.xlsx"
        create_reference_workbooks.main()
        self.assertTrue(original_file.exists())
        workbook = load_workbook(original_file, read_only=False)
        case_prefixes = {name[:12] for name in workbook.sheetnames if name.startswith("CASE-ODE-")}
        self.assertEqual(len(case_prefixes), 12)


if __name__ == "__main__":
    unittest.main()
