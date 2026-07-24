from __future__ import annotations

import json
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from graph_ode.cli import run


class CliIntegrationTests(unittest.TestCase):
    def test_cli_run_writes_required_artifacts(self) -> None:
        try:
            from openpyxl import Workbook
        except ModuleNotFoundError:
            self.skipTest("openpyxl non installato")

        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            workbook_path = tmp_path / "ode.xlsx"
            output_dir = tmp_path / "out"
            wb = Workbook()
            ws = wb.active
            ws.title = "Caso ODE"
            ws["A1"] = "Equazione"
            ws["B1"] = "y' = a*y"
            ws["A2"] = "Parametro"
            ws["B2"] = "a=2"
            ws["A3"] = "Condizione iniziale"
            ws["B3"] = "y(0)=1"
            wb.save(workbook_path)

            code = run(
                [
                    "--input",
                    str(workbook_path),
                    "--output",
                    str(output_dir),
                    "--no-ollama",
                    "--no-pause",
                ]
            )

            self.assertEqual(code, 0)
            required = [
                "workbook_scan.json",
                "candidate_blocks.json",
                "interpretations.json",
                "selected_interpretation.json",
                "solve_result.json",
                "run_report.json",
                "report.md",
                "plot.png",
            ]
            for name in required:
                self.assertTrue((output_dir / name).exists(), name)
            solve_result = json.loads((output_dir / "solve_result.json").read_text(encoding="utf-8"))
            self.assertEqual(solve_result["status"], "solved")
            self.assertIn("exp(2*x)", solve_result["solution"])


if __name__ == "__main__":
    unittest.main()
