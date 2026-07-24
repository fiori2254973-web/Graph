#!/usr/bin/env python3
"""Crea un corpus Excel iniziale allineato ai casi CASE-ODE-001..012."""

from __future__ import annotations

from pathlib import Path


def _require_openpyxl():
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise SystemExit("Installa le dipendenze con: python -m pip install -r SDD_APP\\requirements.txt") from exc
    return Workbook


def main() -> None:
    Workbook = _require_openpyxl()
    output = Path(__file__).resolve().parent / "reference_cases.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    cases = [
        ("CASE-ODE-001", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A2", "Parametro"), ("B2", "a=2"), ("A3", "Condizione iniziale"), ("B3", "y(0)=1")]),
        ("CASE-ODE-002", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A3", "Condizione iniziale"), ("B3", "y(0)=1")]),
        ("CASE-ODE-003", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A2", "Parametro"), ("B2", "a=2"), ("B3", "a=3"), ("A4", "Condizione iniziale"), ("B4", "y(0)=1")]),
        ("CASE-ODE-004", [("A1", "ODE"), ("B1", "y'' + y = 0"), ("A2", "initial"), ("B2", "y(0)=0"), ("B3", "y'(0)=1")]),
        ("CASE-ODE-005", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A2", "Nota"), ("B2", "esempio: a=99 non usare"), ("A5", "Parametro"), ("B5", "a=2"), ("A6", "Condizione iniziale"), ("B6", "y(0)=1")]),
        ("CASE-ODE-006", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A2", "Parametro"), ("B2", "a=2"), ("A3", "Condizione iniziale"), ("B3", "y(0)=1"), ("A5", "Scenario"), ("B5", "LLM contradiction simulated by test_llm_guard.py")]),
        ("CASE-ODE-007-A", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A2", "Parametro"), ("B2", "a=1"), ("A3", "Condizione iniziale"), ("B3", "y(0)=1")]),
        ("CASE-ODE-007-B", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A2", "Parametro"), ("B2", "a=3"), ("A3", "Condizione iniziale"), ("B3", "y(0)=1")]),
        ("CASE-ODE-008", [("A1", "Equazione"), ("B1", "y' = y"), ("A2", "Condizione iniziale"), ("B2", "y(0)=1")]),
        ("CASE-ODE-009", [("A1", "Equazione"), ("B1", "y' = a*y"), ("A2", "Parametro formula"), ("B2", "=1+1"), ("A3", "Condizione iniziale"), ("B3", "y(0)=1")]),
        ("CASE-ODE-010", [("A1", "Equazione"), ("B1", "y' = y"), ("A2", "Condizione iniziale"), ("B2", "y(0)=1"), ("A4", "plot"), ("B4", "x_min=10"), ("B5", "x_max=0")]),
        ("CASE-ODE-011", [("A1", "Equazione ambigua"), ("B1", "dy/dt = a*y"), ("A2", "Parametro"), ("B2", "a=2"), ("A3", "Condizione iniziale"), ("B3", "y(0)=1")]),
        ("CASE-ODE-012", [("A1", "Nota"), ("B1", "l'apostrofo in questo testo non indica y prime"), ("A3", "Equazione"), ("B3", "y' = y"), ("A4", "Condizione iniziale"), ("B4", "y(0)=1")]),
    ]

    for title, cells in cases:
        ws = wb.create_sheet(title)
        for address, value in cells:
            ws[address] = value
        if title == "CASE-ODE-008":
            ws.sheet_state = "hidden"

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
