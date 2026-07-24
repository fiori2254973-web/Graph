#!/usr/bin/env python3
"""Crea workbook Excel di esempio per Graph ODE SDD_APP."""

from __future__ import annotations

from pathlib import Path


def main() -> None:
    try:
        from openpyxl import Workbook
    except ModuleNotFoundError as exc:
        raise SystemExit("Installa le dipendenze con: python -m pip install -r SDD_APP\\requirements.txt") from exc

    base = Path(__file__).resolve().parent
    output = base / "sample_ode.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Caso ODE"
    ws["A1"] = "Equazione"
    ws["B1"] = "y' = a*y"
    ws["A2"] = "Parametro"
    ws["B2"] = "a=2"
    ws["A3"] = "Condizione iniziale"
    ws["B3"] = "y(0)=1"
    ws["A5"] = "Nota"
    ws["B5"] = "esempio descrittivo: non usare come parametro"

    ws2 = wb.create_sheet("Secondo ordine")
    ws2["A1"] = "ODE"
    ws2["B1"] = "y'' + y = 0"
    ws2["A2"] = "initial"
    ws2["B2"] = "y(0)=0"
    ws2["B3"] = "y'(0)=1"

    wb.save(output)
    print(output)


if __name__ == "__main__":
    main()
