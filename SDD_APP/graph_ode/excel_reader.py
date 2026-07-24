from __future__ import annotations

from pathlib import Path
from typing import Any

from .config import AppConfig
from .models import CellRef, WorkbookScan


def _require_openpyxl() -> Any:
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "Dipendenza mancante: openpyxl. Installa con `python -m pip install -r SDD_APP\\requirements.txt`."
        ) from exc
    return openpyxl


def _merged_addresses(worksheet: Any) -> set[str]:
    addresses: set[str] = set()
    for cell_range in worksheet.merged_cells.ranges:
        for row in worksheet[cell_range.coord]:
            for cell in row:
                addresses.add(cell.coordinate)
    return addresses


def scan_workbook(path: Path, config: AppConfig, selected_sheets: list[str] | None = None) -> WorkbookScan:
    if not path.exists():
        return WorkbookScan(
            input_path=str(path),
            sheets_scanned=[],
            cells_seen=0,
            limits=config.as_dict(),
            stop_reason=f"File non trovato: {path}",
        )
    if path.suffix.lower() == ".xls":
        return WorkbookScan(
            input_path=str(path),
            sheets_scanned=[],
            cells_seen=0,
            limits=config.as_dict(),
            stop_reason="Formato .xls legacy non supportato nel rilascio locale v0.1: convertire in .xlsx.",
        )
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return WorkbookScan(
            input_path=str(path),
            sheets_scanned=[],
            cells_seen=0,
            limits=config.as_dict(),
            stop_reason="Formato non supportato: usare .xlsx o .xlsm.",
        )

    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    wanted = set(selected_sheets or workbook.sheetnames)
    warnings: list[str] = []
    scanned: list[str] = []
    cells: list[CellRef] = []
    cells_seen = 0

    if len(workbook.sheetnames) > config.scan_max_sheets:
        return WorkbookScan(
            input_path=str(path),
            sheets_scanned=[],
            cells_seen=0,
            limits=config.as_dict(),
            stop_reason=f"Numero fogli superiore al limite scan_max_sheets={config.scan_max_sheets}.",
        )

    for worksheet in workbook.worksheets:
        if worksheet.title not in wanted:
            continue
        hidden_sheet = worksheet.sheet_state != "visible"
        if hidden_sheet and not config.allow_hidden_sheets:
            warnings.append(f"Foglio nascosto ignorato: {worksheet.title}")
            continue

        scanned.append(worksheet.title)
        merged = _merged_addresses(worksheet)
        hidden_rows = {idx for idx, dim in worksheet.row_dimensions.items() if dim.hidden}
        hidden_cols = {idx for idx, dim in worksheet.column_dimensions.items() if dim.hidden}

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cells_seen += 1
                if cells_seen > config.scan_max_cells:
                    return WorkbookScan(
                        input_path=str(path),
                        sheets_scanned=scanned,
                        cells_seen=cells_seen,
                        limits=config.as_dict(),
                        cells=cells,
                        warnings=warnings,
                        stop_reason=f"Numero celle non vuote superiore al limite scan_max_cells={config.scan_max_cells}.",
                    )
                is_hidden = hidden_sheet or cell.row in hidden_rows or cell.column_letter in hidden_cols
                raw_value = cell.value
                if isinstance(raw_value, str) and raw_value.startswith("="):
                    formula = raw_value
                    display_value = raw_value
                    if not config.allow_formula_cells:
                        warnings.append(f"Formula ignorata per configurazione: {worksheet.title}!{cell.coordinate}")
                        continue
                else:
                    formula = None
                    display_value = str(raw_value)
                cells.append(
                    CellRef(
                        sheet=worksheet.title,
                        address=cell.coordinate,
                        row=cell.row,
                        column=cell.column,
                        raw_value=raw_value,
                        display_value=display_value,
                        formula=formula,
                        data_type=getattr(cell, "data_type", None),
                        is_merged=cell.coordinate in merged,
                        is_hidden=is_hidden,
                    )
                )
                if is_hidden:
                    warnings.append(f"Cella nascosta inclusa come warning: {worksheet.title}!{cell.coordinate}")

    missing = sorted(wanted - set(workbook.sheetnames))
    for name in missing:
        warnings.append(f"Foglio richiesto non presente: {name}")

    return WorkbookScan(
        input_path=str(path),
        sheets_scanned=scanned,
        cells_seen=cells_seen,
        limits=config.as_dict(),
        cells=cells,
        warnings=warnings,
    )
